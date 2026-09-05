"""Create the manuscript figures from results.xlsx."""

import csv
from pathlib import Path

import matplotlib as mpl
mpl.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "data/results.xlsx"
OUT = ROOT / "figures"


def rows(sheet):
    values = list(sheet.iter_rows(values_only=True))
    header = values[0]
    return [dict(zip(header, row)) for row in values[1:] if row[0] is not None]


def csv_rows(path):
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def save(fig, name):
    OUT.mkdir(exist_ok=True)
    svg = OUT / f"{name}.svg"
    fig.savefig(svg, bbox_inches="tight", facecolor="white", metadata={"Date": None})
    clean = "\n".join(line.rstrip() for line in svg.read_text(encoding="utf-8").splitlines())
    svg.write_text(clean + "\n", encoding="utf-8")
    fig.savefig(OUT / f"{name}.png", dpi=600, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  figures/{name}.svg")
    print(f"  figures/{name}.png")


def lengths(herbal_counts, fresh=False):
    food_counts = {}
    source = ROOT / ("work/food/composition_size_distribution.csv" if fresh else "data/food.csv")
    with source.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            food_counts[int(row["ingredient_count"])] = int(row["unique_compositions"])

    sizes = list(range(2, 20))
    food = [food_counts[size] for size in sizes]
    herbal = [herbal_counts[size] for size in sizes]

    print("Ingredient count | Food before matching | Herbal")
    for size, food_count, herb_count in zip(sizes, food, herbal):
        print(f"{size:>16} | {food_count:>20,} | {herb_count:>6,}")
    print("Totals:", f"food={sum(food):,}", f"herbal={sum(herbal):,}")

    fig, ax = plt.subplots(figsize=(7.2, 3.9))
    width = 0.38
    ax.bar([x - width / 2 for x in sizes], food, width, color="#bdbdbd",
           edgecolor="#555555", linewidth=0.5, label="Food before matching (n=762,996)")
    ax.bar([x + width / 2 for x in sizes], herbal, width, color="#222222",
           edgecolor="#222222", linewidth=0.5, label="Herbal (n=2,009)")
    ax.set_yscale("log")
    ax.set(xlabel="Number of ingredients", ylabel="Unique compositions (log scale)", xticks=sizes)
    ax.legend(frameon=False)
    save(fig, "lengths")


def structure(book, fresh=False):
    if fresh:
        key_names = {
            "unique_items": "Unique ingredients",
            "distinct_pairs": "Distinct pairs",
            "item_frequency_gini": "Ingredient-frequency Gini",
            "pair_frequency_gini": "Pair-frequency Gini",
            "item_hhi": "Ingredient HHI",
            "pair_hhi": "Pair HHI",
            "top_1pct_item_share": "Top 1% ingredient share",
            "top_1pct_pair_share": "Top 1% pair share",
            "max_item_prevalence": "Maximum ingredient prevalence",
            "max_pair_prevalence": "Maximum pair prevalence",
            "singleton_pair_share": "Singleton-pair share",
        }
        data = {}
        for row in csv_rows(ROOT / "work/structure/observed_comparison.csv"):
            name = key_names.get(row["metric"])
            if name:
                data[name] = {
                    "herbal_observed": float(row["herb"]),
                    "food_replicate_mean": float(row["food_mean"]),
                    "food_p2_5": float(row["food_p2_5"]),
                    "food_p97_5": float(row["food_p97_5"]),
                }
    else:
        data = {row["metric"]: row for row in rows(book["Structural Comparison"])}
    groups = (
        (
            "Ingredient use",
            (
                ("Unique ingredients", "Unique ingredients"),
                ("Ingredient-frequency Gini", "Frequency Gini"),
                ("Ingredient HHI", "HHI"),
                ("Top 1% ingredient share", "Top 1% share"),
                ("Maximum ingredient prevalence", "Maximum prevalence"),
            ),
        ),
        (
            "Ingredient pairs",
            (
                ("Distinct pairs", "Distinct pairs"),
                ("Pair-frequency Gini", "Frequency Gini"),
                ("Pair HHI", "HHI"),
                ("Top 1% pair share", "Top 1% share"),
                ("Maximum pair prevalence", "Maximum prevalence"),
                ("Singleton-pair share", "Singleton share"),
            ),
        ),
    )
    fig, axes = plt.subplots(1, 2, figsize=(8.0, 4.0))
    print("Metric | Herbal | Food mean | Food 95% interval")
    for _, metrics in groups:
        for key, _ in metrics:
            row = data[key]
            print(
                f"{key} | {float(row['herbal_observed']):.6g} | "
                f"{float(row['food_replicate_mean']):.6g} | "
                f"{float(row['food_p2_5']):.6g}–{float(row['food_p97_5']):.6g}"
            )
    for ax, (title, metrics) in zip(axes, groups):
        labels = [label for _, label in metrics]
        values = [
            100 * (data[key]["herbal_observed"] - data[key]["food_replicate_mean"])
            / data[key]["food_replicate_mean"]
            for key, _ in metrics
        ]
        positions = range(len(values))
        ax.barh(positions, values, color="#333333")
        ax.axvline(0, color="#777777", linewidth=0.8)
        ax.set_yticks(list(positions), labels)
        ax.invert_yaxis()
        ax.set_title(title, fontsize=9)
        ax.set_xlabel("Herbal difference from food mean (%)")
    save(fig, "structure")


def models(book, fresh=False):
    data = csv_rows(ROOT / "work/models/summary.csv") if fresh else rows(book["Model Summary"])
    conditions = ("2", "3", "50%", "75%", "N-1")
    model_names = {
        "popularity": "Popularity",
        "mean_conditional": "Mean conditional probability",
        "mean_jaccard": "Mean pairwise Jaccard",
    }
    colors = {
        "popularity": "#9e9e9e",
        "mean_conditional": "#222222",
        "mean_jaccard": "#666666",
    }
    lookup = {(row["domain"], row["condition"], row["model"]): row for row in data}
    print("Domain | Input | Model | Metric | Mean | SD")
    for domain in ("herbal", "food"):
        for condition in conditions:
            for model in model_names:
                row = lookup[domain, condition, model]
                print(
                    f"{domain} | {condition} | {model} | {row['metric']} | "
                    f"{float(row['mean_performance']):.4f} | {float(row['sd_performance']):.4f}"
                )
    fig, axes = plt.subplots(1, 2, figsize=(8.0, 3.8), sharey=True)
    for ax, domain in zip(axes, ("herbal", "food")):
        for model in model_names:
            selected = [lookup[domain, condition, model] for condition in conditions]
            values = [float(row["mean_performance"]) for row in selected]
            errors = [float(row["sd_performance"]) for row in selected]
            ax.errorbar(conditions, values, yerr=errors, marker="o", linewidth=1.5,
                        capsize=2, color=colors[model], label=model_names[model])
        ax.set_title(domain.title(), fontsize=9)
        ax.set_xlabel("Observed ingredients")
        ax.set_ylim(0.2, 0.6)
    axes[0].set_ylabel("Performance@10")
    axes[1].legend(frameon=False, fontsize=7.5)
    save(fig, "models")


def learning(book, fresh=False):
    data = csv_rows(ROOT / "work/learning/summary.csv") if fresh else rows(book["Learning Curve"])
    print("Domain | Training scale | Mean records | Hit@10 | SD")
    for row in data:
        print(
            f"{row['domain']} | {float(row['training_source_record_scale']):g} | "
            f"{float(row['mean_train_source_records']):.1f} | "
            f"{float(row['mean_hit@10']):.4f} | {float(row['sd_hit@10']):.4f}"
        )
    fig, ax = plt.subplots(figsize=(6.0, 3.9))
    for domain, color, marker in (("herbal", "#222222", "o"), ("food", "#999999", "s")):
        selected = sorted(
            (row for row in data if row["domain"] == domain),
            key=lambda row: float(row["mean_train_source_records"]),
        )
        x = [float(row["mean_train_source_records"]) for row in selected]
        y = [float(row["mean_hit@10"]) for row in selected]
        error = [float(row["sd_hit@10"]) for row in selected]
        ax.errorbar(x, y, yerr=error, marker=marker, linewidth=1.5, capsize=2,
                    color=color, label=domain.title())
    ax.set_xscale("log")
    ax.set(xlabel="Training records (log scale)", ylabel="Hit@10")
    ax.legend(frameon=False)
    save(fig, "learning")


def make_all(herbal_counts, fresh=False):
    mpl.rcParams.update({
        "font.family": "sans-serif",
        "font.sans-serif": ["Arial", "Helvetica", "DejaVu Sans"],
        "font.size": 8.5,
        "axes.linewidth": 0.8,
        "axes.spines.top": False,
        "axes.spines.right": False,
        "axes.grid": True,
        "axes.axisbelow": True,
        "grid.color": "#dddddd",
        "grid.linewidth": 0.6,
        "svg.hashsalt": "HerbalFormulaCompletion",
        "svg.fonttype": "none",
    })
    book = load_workbook(BOOK, read_only=True, data_only=True)
    print("\nFigure 1: composition lengths")
    lengths(herbal_counts, fresh)
    print("\nFigure 2: composition structure")
    structure(book, fresh)
    print("\nFigure 3: recommendation performance")
    models(book, fresh)
    print("\nFigure 4: learning curves")
    learning(book, fresh)
