"""Reproduce the manuscript results and figures."""

import argparse
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from src.figures import make_all
from src.food import paired_records
from src.herbs import open_csv, preprocess


ROOT = Path(__file__).resolve().parent


def title(text):
    print(f"\n{text}\n{'-' * len(text)}")


def show_herbal_raw():
    files = sorted((ROOT / "data/herbal").glob("*.csv"))
    print("Files:", ", ".join(path.name for path in files))
    handle, reader, encoding = open_csv(files[0])
    try:
        first = next(reader)
        print("Header:", ", ".join(reader.fieldnames or []))
        print("First row:", {key: first[key] for key in
              ("처방아이디", "처방한글명", "출전", "약재한글명", "용량", "단위")})
        print("Encoding:", encoding)
    finally:
        handle.close()


def show_food_example():
    example = json.loads((ROOT / "data/example.json").read_text(encoding="utf-8"))
    print("Source: https://im2recipe.csail.mit.edu/")
    print("Access: registration required")
    print("Files used: layer1.json, det_ingrs.json")
    print("Shared ID:", example["id"])
    print("layer1.json fields:", ", ".join(example["layer1"]["fields"]))
    print("det_ingrs.json fields:", ", ".join(example["det_ingrs"]["fields"]))
    print("Raw ingredients:", [item["text"] for item in example["layer1"]["ingredient_excerpt"]])
    print("Detected ingredients:", [item["text"] for item in example["det_ingrs"]["ingredient_excerpt"]])


def prepare_recipe1m():
    data = ROOT / "data/recipe1m"
    archive = data / "recipe1M_layers.tar.gz"
    layer = data / "layer1.json"
    detections = data / "det_ingrs.json"

    if not detections.exists():
        raise SystemExit(f"Missing Recipe1M file: {detections.relative_to(ROOT)}")
    layers = archive if archive.exists() else layer
    if not layers.exists():
        raise SystemExit(
            "Missing Recipe1M file: data/recipe1m/recipe1M_layers.tar.gz "
            "or data/recipe1m/layer1.json"
        )
    print("Reading layer1.json directly from:", layers.name)
    layer_record, detection_record = next(paired_records(layers, detections))
    print("Actual Recipe1M record ID:", layer_record["id"])
    print("layer1.json fields:", ", ".join(layer_record))
    print("det_ingrs.json fields:", ", ".join(detection_record))
    print("Recipe title:", layer_record["title"])
    print("Raw ingredient example:", layer_record["ingredients"][:3])
    print("Detected ingredient example:", detection_record["ingredients"][:3])
    return layers


def show_results():
    book = load_workbook(ROOT / "data/results.xlsx", read_only=True, data_only=True)
    print("Sheets:", ", ".join(book.sheetnames))
    sheet = book["Model Summary"]
    rows = list(sheet.iter_rows(values_only=True))
    print("Recommendation models:", ", ".join(sorted({row[2] for row in rows[1:]})))
    print("Model summary rows:", len(rows) - 1)
    replicates = sum(
        1 for row in book["Structure Replicates"].iter_rows(values_only=True)
        if row[0] is not None
    ) - 1
    print("Food samples:", replicates)
    book.close()


def full_analysis():
    layers = prepare_recipe1m()
    commands = (
        ("src.food", "--layers", str(layers)),
        ("src.structure", "--replicates", "100"),
        ("src.models", "--food-replicates", "100"),
        ("src.learning",),
    )
    for command in commands:
        print("\n$", sys.executable, "-m", *command)
        subprocess.run([sys.executable, "-m", *command], cwd=ROOT, check=True)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--full", action="store_true",
                        help="repeat Recipe1M preprocessing and 100 food samples")
    args = parser.parse_args()

    title("Herbal source data")
    show_herbal_raw()

    title("Herbal preprocessing")
    metadata, collapsed, _, counts = preprocess(
        ROOT / "data/herbal", ROOT / "work/herbal"
    )
    print("Formulas:", f"{metadata['source_formulas']:,}")
    print("Unique compositions:", f"{metadata['unique_compositions']:,}")
    print("Compositions with 2–19 herbs:", f"{metadata['selected_unique_compositions']:,}")
    example = next(row for row in collapsed if "육미지황" in row["formula_names"] and row["weight"] == 19)
    print("Example names:", example["formula_names"])
    print("Example composition:", example["herbs"])
    print("Weight:", example["weight"])

    title("Food source data")
    show_food_example()

    if args.full:
        title("Full Recipe1M analysis")
        full_analysis()

    title("Manuscript results")
    show_results()

    title("Figures")
    make_all(counts, fresh=args.full)


if __name__ == "__main__":
    main()
